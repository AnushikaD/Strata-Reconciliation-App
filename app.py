from __future__ import annotations

import csv
import io
import os
import pickle
import re
import tempfile
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

import pdfplumber
from flask import Flask, render_template, request, send_file
from openpyxl import Workbook, load_workbook


APP_ROOT = Path(__file__).parent
RUNTIME_ROOT = Path(tempfile.gettempdir()) if os.environ.get("VERCEL") else APP_ROOT
UPLOAD_ROOT = RUNTIME_ROOT / "strata_uploads"
UPLOAD_ROOT.mkdir(exist_ok=True)
CACHE_PATH = RUNTIME_ROOT / "strata_last_result.pkl"

app = Flask(__name__, static_folder="public/static", static_url_path="/static")
app.config["MAX_CONTENT_LENGTH"] = 80 * 1024 * 1024


MONEY_RE = re.compile(r"\(?\$?[\d,]+\.\d{2}\)?(?:\s*CR)?")
DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{2}\s+\d+\s+")


def money(value: str | None) -> Decimal:
    if not value:
        return Decimal("0.00")
    raw = str(value).replace("$", "").replace(",", "").strip()
    is_credit = "CR" in raw.upper()
    raw = raw.upper().replace("CR", "").strip()
    is_negative = raw.startswith("(") and raw.endswith(")")
    raw = raw.replace("(", "").replace(")", "")
    amount = Decimal(raw or "0").quantize(Decimal("0.01"))
    return -amount if is_credit or is_negative else amount


def fmt(amount: Decimal | float | int) -> str:
    return f"{Decimal(amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP):,.2f}"


def pdf_text(path: str) -> str:
    with pdfplumber.open(path) as pdf:
        return "\n".join(page.extract_text(x_tolerance=1, y_tolerance=3) or "" for page in pdf.pages)


def read_tabular_text(path: str) -> str:
    suffix = Path(path).suffix.lower()
    if suffix == ".pdf":
        return pdf_text(path)
    if suffix in {".xlsx", ".xls"}:
        wb = load_workbook(path, data_only=True)
        parts = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                parts.append(" ".join("" if v is None else str(v) for v in row))
        return "\n".join(parts)
    if suffix == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as handle:
            return "\n".join(" ".join(row) for row in csv.reader(handle))
    return ""


def parse_balance_sheet(path: str) -> dict[str, Decimal]:
    text = read_tabular_text(path)
    totals = {
        "admin_receivable": Decimal("0.00"),
        "sinking_receivable": Decimal("0.00"),
        "total_receivable": Decimal("0.00"),
        "admin_advance": Decimal("0.00"),
        "sinking_advance": Decimal("0.00"),
        "total_advance": Decimal("0.00"),
        "hvac_advance": Decimal("0.00"),
    }
    for line in text.splitlines():
        values = MONEY_RE.findall(line)
        if "Levies Receivable" in line and len(values) >= 3:
            totals["admin_receivable"], totals["sinking_receivable"], totals["total_receivable"] = map(money, values[-3:])
        elif "Levies Paid in Advance" in line and len(values) >= 3:
            totals["admin_advance"], totals["sinking_advance"], totals["total_advance"] = [abs(money(v)) for v in values[-3:]]
        elif "HVAC/WATERPROOFING" in line or "ADVANCE" in line and "HVAC" in line:
            if len(values) >= 3:
                totals["hvac_advance"] = abs(money(values[-1]))
    return totals


def full_date(short_date: str) -> str:
    return datetime.strptime(short_date, "%d/%m/%y").strftime("%d/%m/%Y")


def parse_owner_transactions(path: str | None = None, text: str | None = None) -> list[dict[str, Any]]:
    if text is None:
        if path is None:
            return []
        text = read_tabular_text(path)
    owner_lookup = parse_owner_lookup(text)
    rows: list[dict[str, Any]] = []
    lot = unit = owner = ""
    current: dict[str, Any] | None = None
    status = ""

    def flush_description(extra: str) -> None:
        nonlocal current
        if current:
            current["description"] = f"{current['description']} {extra}".strip()

    for raw in text.splitlines():
        line = raw.replace("(cid:9)", " ").strip()
        lot_match = re.search(r"Lot#:\s*([A-Z0-9]+).*?Unit#:\s*([A-Z0-9]+)", line)
        if not lot_match:
            lot_match = re.search(r"Lot#:\s*([A-Z0-9]+).*?(\d{3})\s+Units of Entitlement", line)
        if lot_match:
            lot, unit = lot_match.groups()
            if lot in owner_lookup:
                unit = owner_lookup[lot]["unit"]
                owner = owner_lookup[lot]["owner"]
            current = None
            status = ""
            continue
        owner_match = re.search(r"Owner Name:\s*(.*?)(?:\s+Contribution Schedule:|\s+Interest \(Insurance\)|$)", line)
        if owner_match:
            owner = owner_match.group(1).strip()
            continue
        if DATE_RE.match(line):
            values = MONEY_RE.findall(line)
            if len(values) >= 7:
                prefix = MONEY_RE.split(line, maxsplit=1)[0]
                due, ref, desc = prefix.split(maxsplit=2)
                current = {"due_date": full_date(due), "reference": ref, "description": desc.strip(), "lot": lot, "unit": unit, "owner": owner}
                status = ""
            continue
        if current and not line.startswith(("FULLY PAID", "OVERDUE", "Admin ", "Sinking ", "Levy Totals", "Receipts", "Date Reference")):
            if not any(skip in line for skip in ["Due Date Reference Details", "Contribution Schedule", "TCM AGREEMENTS"]):
                flush_description(line)
            continue
        fund_line = line
        if line.startswith(("FULLY PAID ", "OVERDUE ")):
            status, fund_line = line.split(maxsplit=1)
        if current and fund_line.startswith(("Admin ", "Sinking ")):
            values = MONEY_RE.findall(fund_line)
            if len(values) >= 7 and status == "OVERDUE":
                fund = "Administrative Fund" if fund_line.startswith("Admin ") else "Sinking Fund"
                unpaid = money(values[3])
                if unpaid > 0:
                    rows.append({
                        "lot": current["lot"],
                        "unit": current["unit"],
                        "owner": current["owner"],
                        "due_date": current["due_date"],
                        "description": current["description"],
                        "fund": fund,
                        "amount": unpaid,
                        "type": "Arrear",
                        "confidence": 0.98,
                        "notes": "Matched overdue split line in Owner Transaction Summary.",
                    })
    return sorted(rows, key=lambda r: (str(r["lot"]).zfill(5), r["fund"], datetime.strptime(r["due_date"], "%d/%m/%Y")))


def parse_owner_lookup(text: str) -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    lot = unit = owner = ""
    awaiting_owner_tail = False
    for raw in text.splitlines():
        line = raw.replace("(cid:9)", " ").strip()
        lot_match = re.search(r"Lot#:\s*([A-Z0-9]+).*?Unit#:\s*([A-Z0-9]+)", line)
        if not lot_match:
            lot_match = re.search(r"Lot#:\s*([A-Z0-9]+).*?(\d{3})\s+Units of Entitlement", line)
        if lot_match:
            lot, unit = lot_match.groups()
            awaiting_owner_tail = False
            continue
        owner_match = re.search(r"Owner Name:\s*(.*?)(?:\s+Contribution Schedule:|\s+Interest \(Insurance\)|$)", line)
        if owner_match and lot:
            owner = owner_match.group(1).strip()
            awaiting_owner_tail = True
            lookup[lot] = {"unit": unit, "owner": owner}
            continue
        if awaiting_owner_tail and lot and "Interest (Insurance)" in line:
            tail = line.split("Interest (Insurance)", 1)[0].strip()
            if tail:
                lookup[lot]["owner"] = f"{lookup[lot]['owner']} {tail}".strip()
            awaiting_owner_tail = False
    return lookup


def parse_lot_arrears_meta(text: str) -> dict[str, dict[str, Decimal]]:
    meta: dict[str, dict[str, Decimal]] = {}
    lot = ""
    for raw in text.splitlines():
        line = raw.replace("(cid:9)", " ").strip()
        lot_match = re.search(r"Lot#:\s*([A-Z0-9]+).*?(?:Unit#:\s*[A-Z0-9]+|(\d{3})\s+Units of Entitlement)", line)
        if lot_match:
            lot = lot_match.group(1)
            arrears_match = re.search(r"Arrears:\s*(\(?\$?[\d,]+\.\d{2}\)?(?:\s*CR)?)", line)
            meta.setdefault(lot, {"header_arrears": Decimal("0.00"), "period_unpaid": Decimal("0.00")})
            if arrears_match:
                meta[lot]["header_arrears"] = abs(money(arrears_match.group(1)))
            continue
        if lot and "Arrears:" in line:
            arrears_match = re.search(r"Arrears:\s*(\(?\$?[\d,]+\.\d{2}\)?(?:\s*CR)?)", line)
            if arrears_match:
                meta.setdefault(lot, {"header_arrears": Decimal("0.00"), "period_unpaid": Decimal("0.00")})
                meta[lot]["header_arrears"] = abs(money(arrears_match.group(1)))
        if lot and line.startswith("Levy Totals"):
            values = MONEY_RE.findall(line)
            if len(values) >= 4:
                meta.setdefault(lot, {"header_arrears": Decimal("0.00"), "period_unpaid": Decimal("0.00")})
                meta[lot]["period_unpaid"] = abs(money(values[3]))
    return meta


def parse_position_line(line: str) -> dict[str, Any] | None:
    values = MONEY_RE.findall(line)
    if len(values) < 6:
        return None
    first_amount = MONEY_RE.search(line)
    if not first_amount:
        return None
    prefix = line[: first_amount.start()].strip()
    parts = prefix.split(maxsplit=2)
    if len(parts) < 2:
        return None
    lot, unit = parts[0], parts[1]
    owner = parts[2] if len(parts) > 2 else ""
    return {"lot": lot, "unit": unit, "owner": owner, "values": values}


def parse_levy_positions(
    path: str,
    owner_lookup: dict[str, dict[str, str]] | None = None,
    arrears_meta: dict[str, dict[str, Decimal]] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if Path(path).suffix.lower() != ".pdf":
        return [], []
    owner_lookup = owner_lookup or {}
    arrears_meta = arrears_meta or {}
    advances: list[dict[str, Any]] = []
    opening_arrears: list[dict[str, Any]] = []
    seen_advances: set[tuple[str, str, str, Decimal]] = set()
    seen_opening: set[tuple[str, str, Decimal]] = set()
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            fund = "Administrative Fund" if "Administrative Fund" in text else "Sinking Fund" if "Sinking Fund" in text else ""
            for line in text.splitlines():
                parsed = parse_position_line(line.strip())
                if not parsed:
                    continue
                lot = parsed["lot"]
                values = parsed["values"]
                unit = owner_lookup.get(lot, {}).get("unit", parsed["unit"])
                owner = owner_lookup.get(lot, {}).get("owner", parsed["owner"])
                opening = money(values[0])
                closing = money(values[4])
                if "CR" not in values[4].upper() and closing > 0:
                    pass
                if "CR" in values[4].upper() and abs(closing) > 0:
                    key = (lot, unit, fund, abs(closing))
                    if key not in seen_advances:
                        seen_advances.add(key)
                        advances.append({
                            "lot": lot,
                            "unit": unit,
                            "owner": owner,
                            "fund": fund,
                            "amount": abs(closing),
                            "confidence": 0.9 if lot in owner_lookup else 0.78,
                            "notes": "Credit closing balance in Levy Positions report.",
                        })
                opening_needed = arrears_meta.get(lot, {}).get("header_arrears", Decimal("0.00")) - arrears_meta.get(lot, {}).get("period_unpaid", Decimal("0.00"))
                if "CR" not in values[0].upper() and opening > 0 and opening_needed > Decimal("0.005"):
                    key = (lot, fund, opening)
                    if key not in seen_opening:
                        seen_opening.add(key)
                        opening_arrears.append({
                            "lot": lot,
                            "unit": unit,
                            "owner": owner,
                            "due_date": "01/05/2025",
                            "description": "Opening Balance",
                            "fund": fund,
                            "amount": opening,
                            "type": "Arrear",
                            "confidence": 0.88 if lot in owner_lookup else 0.76,
                            "notes": "Positive opening balance in Levy Positions report.",
                        })
    advances.sort(key=lambda r: (str(r["lot"]).zfill(5), r["fund"], -r["amount"]))
    opening_arrears.sort(key=lambda r: (str(r["lot"]).zfill(5), r["fund"]))
    return advances, opening_arrears


def totals_by_fund(rows: list[dict[str, Any]]) -> dict[str, Decimal]:
    result = {"Administrative Fund": Decimal("0.00"), "Sinking Fund": Decimal("0.00")}
    for row in rows:
        if row.get("fund") in result:
            result[row["fund"]] += row["amount"]
    return result


def reconcile(balance_path: str, positions_path: str, owner_path: str) -> dict[str, Any]:
    balance = parse_balance_sheet(balance_path)
    owner_text = read_tabular_text(owner_path)
    owner_lookup = parse_owner_lookup(owner_text)
    arrears_meta = parse_lot_arrears_meta(owner_text)
    arrears = parse_owner_transactions(text=owner_text)
    advances, opening_arrears = parse_levy_positions(positions_path, owner_lookup, arrears_meta)
    existing = {(r["lot"], r["due_date"], r["fund"], r["description"], r["amount"]) for r in arrears}
    arrears.extend(r for r in opening_arrears if (r["lot"], r["due_date"], r["fund"], r["description"], r["amount"]) not in existing)
    arrears.sort(key=lambda r: (str(r["lot"]).zfill(5), r["fund"], datetime.strptime(r["due_date"], "%d/%m/%Y")))
    arrears_total = sum((r["amount"] for r in arrears), Decimal("0.00"))
    advances_total = sum((r["amount"] for r in advances), Decimal("0.00"))
    arrears_funds = totals_by_fund(arrears)
    advances_funds = totals_by_fund(advances)
    return {
        "balance": balance,
        "arrears": arrears,
        "advances": advances,
        "summary": {
            "arrears_total": arrears_total,
            "advances_total": advances_total,
            "admin_arrears": arrears_funds["Administrative Fund"],
            "sinking_arrears": arrears_funds["Sinking Fund"],
            "admin_advances": advances_funds["Administrative Fund"],
            "sinking_advances": advances_funds["Sinking Fund"],
            "arrears_variance": arrears_total - balance["total_receivable"],
            "admin_arrears_variance": arrears_funds["Administrative Fund"] - balance["admin_receivable"],
            "sinking_arrears_variance": arrears_funds["Sinking Fund"] - balance["sinking_receivable"],
            "advances_variance": advances_total - balance["total_advance"],
        },
    }


def save_upload(name: str) -> str:
    file = request.files[name]
    target = UPLOAD_ROOT / file.filename
    file.save(target)
    return str(target)


LAST_RESULT: dict[str, Any] | None = None


def save_last_result(result: dict[str, Any]) -> None:
    with CACHE_PATH.open("wb") as handle:
        pickle.dump(result, handle)


def get_last_result() -> dict[str, Any] | None:
    global LAST_RESULT
    if LAST_RESULT:
        return LAST_RESULT
    if CACHE_PATH.exists():
        with CACHE_PATH.open("rb") as handle:
            LAST_RESULT = pickle.load(handle)
    return LAST_RESULT


@app.route("/", methods=["GET", "POST"])
def index():
    global LAST_RESULT
    result = None
    if request.method == "POST":
        paths = {key: save_upload(key) for key in ["balance_sheet", "levy_positions", "owner_summary"]}
        result = reconcile(paths["balance_sheet"], paths["levy_positions"], paths["owner_summary"])
        LAST_RESULT = result
        save_last_result(result)
    return render_template("index.html", result=result, fmt=fmt)


def export_rows(kind: str) -> tuple[list[str], list[list[Any]]]:
    result = get_last_result()
    if not result:
        return [], []
    if kind == "arrears":
        headers = ["Lot #", "Unit #", "Owner", "Due Date", "Description", "Fund", "Amount", "Type"]
        rows = [[r["lot"], r["unit"], r["owner"], r["due_date"], r["description"], r["fund"], float(r["amount"]), r["type"]] for r in result["arrears"]]
    else:
        headers = ["Lot #", "Unit #", "Owner", "Fund", "Amount"]
        rows = [[r["lot"], r["unit"], r["owner"], r["fund"], float(r["amount"])] for r in result["advances"]]
    return headers, rows


@app.route("/export/<kind>.xlsx")
def export(kind: str):
    headers, rows = export_rows(kind)
    wb = Workbook()
    ws = wb.active
    ws.title = "Prepaid" if kind == "advances" else kind.title()
    ws.append(headers)
    for row in rows:
        ws.append(row)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    wb.save(tmp.name)
    return send_file(tmp.name, as_attachment=True, download_name=f"{kind}.xlsx")


if __name__ == "__main__":
    app.run(host="0.0.0.0", debug=False, port=int(os.environ.get("PORT", "5000")))
